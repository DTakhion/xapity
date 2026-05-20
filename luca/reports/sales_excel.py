# luca/reports/sales_excel.py

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from db.mongo_persistence_luca import (
    find_luca_sales_snapshot,
    insert_luca_report_metadata,
)
from luca.analytics.sales_summary import build_sales_summary_from_snapshot


ROOT_DIR = Path(__file__).resolve().parents[2]
REPORTS_DIR = ROOT_DIR / "luca" / "reports" / "output"


MONTH_NAMES = {
    0: "todos",
    1: "enero",
    2: "febrero",
    3: "marzo",
    4: "abril",
    5: "mayo",
    6: "junio",
    7: "julio",
    8: "agosto",
    9: "septiembre",
    10: "octubre",
    11: "noviembre",
    12: "diciembre",
}


def safe_float(value: Any) -> float:
    return float(value or 0)


def get_libros(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    return snapshot.get("data", {}).get("libros", [])


def classify_financial_status(item: dict[str, Any]) -> str:
    status = str(item.get("status", "")).upper()
    due_status = str(item.get("diasHastaVencimiento", "")).lower()

    if status == "COBRADO":
        return "COBRADO"

    if due_status == "vencida":
        return "PENDIENTE VENCIDO"

    return "PENDIENTE NO VENCIDO"


def style_header(ws, row: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column

        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[get_column_letter(column)].width = min(
            max(max_length + 2, 12),
            45,
        )


def write_summary_sheet(
    wb: Workbook,
    summary: dict[str, Any],
) -> None:
    ws = wb.active
    ws.title = "Resumen"

    resumen = summary.get("resumen", {})

    ws["A1"] = "Reporte de ventas Luca"
    ws["A1"].font = Font(size=16, bold=True)

    rows_top = [
        ("Business ID", summary.get("businessId")),
        ("Periodo", f"{MONTH_NAMES.get(summary.get('month'))} {summary.get('year')}"),
        ("Tipo", summary.get("type")),
        ("Registros", summary.get("recordsCount")),
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Snapshot origen", summary.get("sourceSnapshotId")),
    ]

    for idx, (label, value) in enumerate(rows_top, start=3):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)

    rows = [
        ("Total documentos", resumen.get("totalDocumentos")),
        ("Monto total ventas", resumen.get("montoTotalVentas")),
        ("Monto cobrado", resumen.get("montoCobrado")),
        ("Monto pendiente", resumen.get("montoPendiente")),
        ("Monto pendiente vencido", resumen.get("montoPendienteVencido")),
        ("Monto IVA retenido", resumen.get("montoIvaRetenido")),
        ("Monto neto líquido", resumen.get("montoNetoLiquido")),
        ("Monto exento", resumen.get("montoExento")),
        ("Cantidad cobrados", resumen.get("cantidadCobrados")),
        ("Cantidad pendientes", resumen.get("cantidadPendientes")),
        ("Cantidad pendientes vencidos", resumen.get("cantidadPendientesVencidos")),
    ]

    start_row = 11
    ws.cell(start_row, 1, "Métrica")
    ws.cell(start_row, 2, "Valor")
    style_header(ws, start_row)

    for idx, (label, value) in enumerate(rows, start=start_row + 1):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)

        if "Monto" in label:
            ws.cell(idx, 2).number_format = '$ #,##0'

    auto_width(ws)


def write_top_clients_sheet(
    wb: Workbook,
    summary: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Top clientes")

    headers = [
        "RUT",
        "Razón social",
        "Cantidad documentos",
        "Monto total",
        "Monto cobrado",
        "Monto pendiente",
    ]

    ws.append(headers)
    style_header(ws, 1)

    for item in summary.get("topClientes", []):
        ws.append([
            item.get("rut"),
            item.get("razonSocial"),
            item.get("cantidadDocumentos"),
            safe_float(item.get("montoTotal")),
            safe_float(item.get("montoCobrado")),
            safe_float(item.get("montoPendiente")),
        ])

    for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '$ #,##0'

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws)


def write_status_sheet(
    wb: Workbook,
    summary: dict[str, Any],
) -> None:
    ws = wb.create_sheet("Estados")

    ws.append(["Status", "Cantidad"])
    style_header(ws, 1)

    for status, count in summary.get("statusCounts", {}).items():
        ws.append([status, count])

    ws.auto_filter.ref = ws.dimensions
    auto_width(ws)


def write_documents_sheet(
    wb: Workbook,
    snapshot: dict[str, Any],
    sheet_name: str,
    documents: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Folio",
        "Fecha emisión",
        "Fecha vencimiento",
        "Fecha pago",
        "RUT",
        "Razón social",
        "Tipo comprobante",
        "Nombre folio",
        "Monto total",
        "Monto IVA retenido",
        "Monto neto líquido",
        "Monto exento",
        "Status Luca",
        "Estado financiero",
        "Score",
        "Tiene linkage",
        "Monto relacionado",
    ]

    ws.append(headers)
    style_header(ws, 1)

    for item in documents:
        linkage = item.get("linkage") or []
        monto_relacionado = sum(
            safe_float(link.get("montoRelacionado"))
            for link in linkage
        )

        ws.append([
            item.get("folio") or item.get("numeroFolio"),
            item.get("fecha"),
            item.get("fechaVencimiento"),
            item.get("fechaPago"),
            item.get("rut"),
            item.get("razonSocial"),
            item.get("tipoComprobante"),
            item.get("nombreFolio"),
            safe_float(item.get("montoTotal")),
            safe_float(item.get("montoIvaRetenido")),
            safe_float(item.get("montoNetoLiquido")),
            safe_float(item.get("montoExento")),
            item.get("status"),
            classify_financial_status(item),
            item.get("score"),
            "Sí" if linkage else "No",
            monto_relacionado,
        ])

    for row in ws.iter_rows(min_row=2, min_col=9, max_col=12):
        for cell in row:
            cell.number_format = '$ #,##0'

    for row in ws.iter_rows(min_row=2, min_col=17, max_col=17):
        for cell in row:
            cell.number_format = '$ #,##0'

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws)


def generate_sales_excel(
    business_id: int,
    month: int,
    year: int,
    type_: str = "ingreso",
    linkage: bool = True,
    requested_by: str | None = None,
) -> Path:
    snapshot = find_luca_sales_snapshot(
        business_id=business_id,
        year=year,
        month=month,
        type_=type_,
        linkage=linkage,
    )

    if not snapshot:
        raise FileNotFoundError(
            f"No existe snapshot Mongo para businessId={business_id}, "
            f"year={year}, month={month}, type={type_}, linkage={linkage}"
        )

    summary = build_sales_summary_from_snapshot(
        snapshot=snapshot,
        requested_by=requested_by,
        persist=True,
    )

    libros = get_libros(snapshot)

    cobrados = [
        item for item in libros
        if str(item.get("status", "")).upper() == "COBRADO"
    ]

    pendientes = [
        item for item in libros
        if str(item.get("status", "")).upper() != "COBRADO"
    ]

    pendientes_vencidos = [
        item for item in pendientes
        if str(item.get("diasHastaVencimiento", "")).lower() == "vencida"
    ]

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    month_label = f"{month:02d}" if month != 0 else "00_todos"

    output_path = (
        REPORTS_DIR
        / f"reporte_ventas_{type_}_business_{business_id}_{year}_{month_label}.xlsx"
    )

    wb = Workbook()

    write_summary_sheet(wb, summary)
    write_top_clients_sheet(wb, summary)
    write_status_sheet(wb, summary)

    write_documents_sheet(
        wb=wb,
        snapshot=snapshot,
        sheet_name="Documentos",
        documents=libros,
    )

    write_documents_sheet(
        wb=wb,
        snapshot=snapshot,
        sheet_name="Cobrados",
        documents=cobrados,
    )

    write_documents_sheet(
        wb=wb,
        snapshot=snapshot,
        sheet_name="Pendientes vencidos",
        documents=pendientes_vencidos,
    )

    write_documents_sheet(
        wb=wb,
        snapshot=snapshot,
        sheet_name="Pendientes",
        documents=pendientes,
    )

    wb.save(output_path)

    report_id = insert_luca_report_metadata({
        "businessId": business_id,
        "year": year,
        "month": month,
        "type": type_,
        "linkage": linkage,
        "reportType": "sales_excel",
        "requestedBy": requested_by,
        "generatedAt": datetime.now(),
        "sourceSnapshotId": summary.get("sourceSnapshotId"),
        "summaryId": summary.get("summaryId"),
        "filePath": str(output_path),
        "storageProvider": "local",
        "fileName": output_path.name,
        "status": "generated",
    })

    print(f"Reporte registrado en Mongo con reportId={report_id}")

    return output_path


if __name__ == "__main__":
    path = generate_sales_excel(
        business_id=11,
        month=1,
        year=2026,
        requested_by="local-test",
    )

    print(f"Excel generado en: {path}")